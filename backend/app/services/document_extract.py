"""
Extrai texto de arquivos enviados pra Central de Contexto (tabela de preços,
política da loja, FAQ, etc.) — o texto extraído é o que entra no prompt dos
agentes de IA, então precisa ser texto puro e razoavelmente enxuto.
"""
import io

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

MAX_FILE_BYTES = 5 * 1024 * 1024  # 5MB
MAX_TEXT_CHARS = 50_000  # por documento, guardado no banco

SUPPORTED_EXTENSIONS = {"pdf", "docx", "xlsx", "csv", "txt", "md"}


class UnsupportedFileError(Exception):
    pass


def extension(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _extract_pdf(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_docx(content: bytes) -> str:
    doc = Document(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_xlsx(content: bytes) -> str:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"## {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def extract_text(filename: str, content: bytes) -> str:
    ext = extension(filename)
    if ext not in SUPPORTED_EXTENSIONS:
        raise UnsupportedFileError(
            f"Formato \".{ext or '?'}\" não suportado. Envie PDF, Word (.docx), "
            "Excel (.xlsx), CSV ou texto (.txt/.md)."
        )

    try:
        if ext in ("txt", "md", "csv"):
            text = content.decode("utf-8", errors="ignore")
        elif ext == "pdf":
            text = _extract_pdf(content)
        elif ext == "docx":
            text = _extract_docx(content)
        else:  # xlsx
            text = _extract_xlsx(content)
    except Exception as e:
        raise UnsupportedFileError(f"Não consegui ler esse arquivo: {e}")

    text = text.strip()
    if not text:
        raise UnsupportedFileError(
            "Não encontrei texto nesse arquivo — ele pode estar vazio, ser uma "
            "imagem escaneada, ou protegido por senha."
        )
    return text[:MAX_TEXT_CHARS]
